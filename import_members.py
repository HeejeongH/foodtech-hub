"""
Import members from CSV or XLSX (exported from the Google Sheet).

Usage (CLI):
    python import_members.py members.csv
    python import_members.py members.xlsx

The expected columns (Korean headers from the sheet):
    연번, 대분류, 소분류, 구분, 세부구분, 성명, 직위, 소속, 소재지, 세부소속,
    휴대폰, 이메일, * 총동문회, * 월드푸드테크협의회, 홈페이지,
    연구분야(대학소속) / 사업분야(기업소속), 비고, ...

Header column mapping:
    소분류       → cohort      (예: '원우(10기)')
    구분         → category    (기업/기관/대학/언론)
    세부구분     → subcategory (일반/스타트업/대학/...)
    성명         → name        (필수)
    * 총동문회   → membership_status
    * 월드푸드테크협의회 → benefit_pct
    유료회원여부 → membership_type
    비고         → notes + payment_history(납입 이력)
"""
import sys, csv, secrets, re, io
from pathlib import Path
from typing import Optional, Iterable

from db import init_db, engine, Member
from sqlmodel import Session, select


# ============================================================
# Cleaners
# ============================================================
def _norm(s) -> Optional[str]:
    if s is None: return None
    s = str(s).strip()
    if s in ("", "-", "—", "X", "x", "None", "nan"): return None
    return s

def _norm_phone(s) -> Optional[str]:
    if not s: return None
    s = re.sub(r"[^\d\-+]", "", str(s))
    return s or None

def _split_email(s) -> Optional[str]:
    if not s: return None
    s = str(s).replace("\n", ",").replace(";", ",")
    parts = [p.strip() for p in s.split(",") if "@" in p]
    return parts[0] if parts else None


def parse_row(row: dict) -> Optional[dict]:
    """Map sheet columns to Member fields. Returns None if row should be skipped."""
    name = _norm(row.get("성명"))
    if not name: return None
    notes = _norm(row.get("비고"))
    payment = None
    if notes and ("납입" in notes or "이후" in notes):
        payment = notes
    return {
        "name": name,
        "email": _split_email(row.get("이메일")),
        "phone": _norm_phone(row.get("휴대폰")),
        "cohort": _norm(row.get("소분류")),
        "category": _norm(row.get("구분")),
        "subcategory": _norm(row.get("세부구분")),
        "position": _norm(row.get("직위")),
        "organization": _norm(row.get("소속")),
        "location": _norm(row.get("소재지")),
        "division": _norm(row.get("세부소속")),
        "business_area": _norm(row.get("연구분야(대학소속) / 사업분야(기업소속)") or row.get("사업분야")),
        "membership_status": _norm(row.get("* 총동문회")) or _norm(row.get("회원 여부 (2025.04.01. 기준)")),
        "membership_type": _norm(row.get("유료회원여부 (26.6.18.)")),
        "payment_history": payment,
        "benefit_pct": _norm(row.get("* 월드푸드테크협의회")) or _norm(row.get("혜택 적용 여부 (수강료 감면)")),
        "council_label": _norm(row.get("소속 (협의회 표기 기준)")),
        "notes": notes,
    }


# ============================================================
# File readers (CSV / XLSX, with header-row autodetection)
# ============================================================
EXPECTED_HEADERS = {"성명", "이메일", "소분류", "구분", "휴대폰"}

def _find_header_row(rows: list, max_check: int = 15) -> int:
    """Some sheets have title/comment rows above the real header.
    Find the row that contains the most expected Korean headers."""
    best = 0; best_score = 0
    for i, row in enumerate(rows[:max_check]):
        score = sum(1 for c in row if c and str(c).strip() in EXPECTED_HEADERS)
        if score > best_score:
            best_score = score; best = i
    return best


def read_csv(content: bytes) -> list[dict]:
    """Read CSV from bytes, autodetect encoding (UTF-8/UTF-8-BOM/CP949/EUC-KR)."""
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            text = content.decode(enc)
            print(f"[csv] decoded as {enc}")
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Cannot decode CSV (tried utf-8, cp949, euc-kr)")
    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)
    if not all_rows: return []
    header_idx = _find_header_row(all_rows)
    print(f"[csv] header row #{header_idx + 1}: {all_rows[header_idx]}")
    headers = [str(h).strip() for h in all_rows[header_idx]]
    out = []
    for r in all_rows[header_idx + 1:]:
        if not any(r): continue
        row_dict = dict(zip(headers, [str(v).strip() if v else "" for v in r]))
        out.append(row_dict)
    return out


def read_xlsx(content: bytes, sheet_name: Optional[str] = None) -> list[dict]:
    """Read XLSX (Excel) from bytes."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    print(f"[xlsx] sheet: {ws.title} ({ws.max_row} rows × {ws.max_column} cols)")
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(["" if v is None else str(v).strip() for v in row])
    if not all_rows: return []
    header_idx = _find_header_row(all_rows)
    headers = all_rows[header_idx]
    print(f"[xlsx] header row #{header_idx + 1}: {[h for h in headers if h]}")
    out = []
    for r in all_rows[header_idx + 1:]:
        if not any(r): continue
        out.append(dict(zip(headers, r)))
    return out


def read_file(path_or_bytes, filename: str = "") -> list[dict]:
    """Auto-detect format and read."""
    if isinstance(path_or_bytes, (bytes, bytearray)):
        content = bytes(path_or_bytes)
    else:
        p = Path(path_or_bytes)
        content = p.read_bytes()
        filename = filename or p.name
    name = filename.lower()
    if name.endswith((".xlsx", ".xlsm", ".xls")):
        return read_xlsx(content)
    return read_csv(content)


# ============================================================
# Upsert
# ============================================================
def upsert(session: Session, data: dict) -> tuple[str, Member]:
    """Returns ('created'|'updated', Member)."""
    existing = None
    if data.get("email"):
        existing = session.exec(select(Member).where(Member.email == data["email"])).first()
    if not existing:
        existing = session.exec(
            select(Member).where(Member.name == data["name"], Member.organization == data.get("organization"))
        ).first()
    if existing:
        for k, v in data.items():
            if v is not None: setattr(existing, k, v)
        if not existing.unsubscribe_token:
            existing.unsubscribe_token = secrets.token_urlsafe(24)
        session.add(existing)
        return "updated", existing
    m = Member(**data, unsubscribe_token=secrets.token_urlsafe(24))
    session.add(m)
    return "created", m


def import_rows(rows: Iterable[dict]) -> dict:
    """Import rows into DB. Returns summary."""
    init_db()
    created = updated = skipped = 0
    errors = []
    with Session(engine) as session:
        for i, row in enumerate(rows, 2):
            try:
                data = parse_row(row)
                if not data:
                    skipped += 1; continue
                action, _ = upsert(session, data)
                if action == "created": created += 1
                else: updated += 1
                if (created + updated) % 50 == 0:
                    session.commit()
            except Exception as e:
                errors.append({"row": i, "error": str(e)})
        session.commit()
    return {"created": created, "updated": updated, "skipped": skipped,
            "errors": errors, "total_processed": created + updated + skipped}


def import_file(path: str) -> dict:
    """CLI entry."""
    p = Path(path)
    if not p.exists():
        print(f"[error] file not found: {path}"); sys.exit(1)
    rows = read_file(p)
    print(f"[import] loaded {len(rows)} data rows")
    result = import_rows(rows)
    print(f"[import] done · created={result['created']} updated={result['updated']} skipped={result['skipped']}")
    if result["errors"]:
        print(f"[import] {len(result['errors'])} row errors:")
        for e in result["errors"][:10]:
            print(f"  row {e['row']}: {e['error']}")
    return result


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python import_members.py <csv-or-xlsx-file>")
        sys.exit(1)
    import_file(sys.argv[1])
